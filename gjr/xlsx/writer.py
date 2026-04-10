"""把翻译结果写回 Excel。

核心函数 `_save_xlsx_with_images` 手写 XML 替换,**不走 openpyxl 保存**:
  - openpyxl 保存时会丢图片、丢部分格式
  - xlsx 本质是 zip,我们直接替换 `xl/sharedStrings.xml` 和 `xl/worksheets/sheet1.xml`
  - 其他所有条目(图片、样式、主题)原封不动拷贝

.xls 走 fallback 路径 xls_to_xlsx(openpyxl 写新文件),会丢失图片。
"""
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from pathlib import Path

import xlrd
from openpyxl import Workbook, load_workbook


def _estimate_text_width(text):
    """粗略估算一段文本的显示宽度(按字符类型加权)。"""
    width = 0
    for ch in str(text):
        if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f':
            width += 2  # CJK
        elif '\u0e00' <= ch <= '\u0e7f':
            width += 1.5  # Thai
        elif '\u1000' <= ch <= '\u109f':
            width += 1.5  # Burmese
        else:
            width += 1
    return width


def xls_to_xlsx(xls_path):
    """.xls → 内存里的新 Workbook(不保存;丢图片)。fallback 路径。"""
    wb_old = xlrd.open_workbook(xls_path, formatting_info=True)
    sheet = wb_old.sheet_by_index(0)
    wb_new = Workbook()
    ws = wb_new.active
    ws.title = sheet.name
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            val = sheet.cell_value(row, col)
            if val:
                ws.cell(row=row+1, column=col+1, value=val)
    for rlo, rhi, clo, chi in sheet.merged_cells:
        ws.merge_cells(start_row=rlo+1, start_column=clo+1, end_row=rhi, end_column=chi)
    return wb_new


def save_translated_xlsx(original_path, translations, output_path):
    """统一入口:按扩展名分派到 .xlsx 的图片保留路径或 .xls 的 fallback。"""
    ext = Path(original_path).suffix.lower()
    if ext == '.xlsx':
        _save_xlsx_with_images(original_path, translations, output_path)
    else:
        wb = xls_to_xlsx(original_path)
        ws = wb.active
        for (row, col), val in translations.items():
            ws.cell(row=row+1, column=col+1, value=val)
        wb.save(output_path)


NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_NS_PREFIX = f'{{{NS}}}'


def _save_xlsx_with_images(original_path, translations, output_path):
    """
    .xlsx 保图片的写入路径。

    流程:
      1. 用 openpyxl 读一次,拿到合并单元格信息(skip_cells 防止把合并单元格的从格也写了)
      2. 直接 unzip 原 .xlsx,提取三个 XML: sharedStrings, sheet1, styles
      3. 修改 XML:
         - 共享字符串表里的条目整体替换
         - sheet1 里 inlineStr/cell 节点的文本替换
         - 列宽按中文长度自动加宽
         - 样式表加上 wrapText(文字太长时换行显示)
      4. zip 重写:改过的三个 XML 替换,其他条目原样复制(图片/主题等)
    """
    # 1. 先用 openpyxl 读合并单元格信息
    wb = load_workbook(original_path)
    ws = wb.active

    skip_cells = set()
    merged_col_spans = {}
    for merged_range in ws.merged_cells.ranges:
        col_span = merged_range.max_col - merged_range.min_col + 1
        merged_col_spans[(merged_range.min_row, merged_range.min_col)] = col_span
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if (row, col) != (merged_range.min_row, merged_range.min_col):
                    skip_cells.add((row, col))
    wb.close()

    ns = {'ns': NS}
    ET.register_namespace('', NS)

    # 2. 从 zip 里读三个 XML
    with zipfile.ZipFile(original_path, 'r') as zin:
        has_shared_strings = 'xl/sharedStrings.xml' in zin.namelist()
        shared_strings_xml = zin.read('xl/sharedStrings.xml') if has_shared_strings else None
        sheet_xml = zin.read('xl/worksheets/sheet1.xml')
        styles_xml = zin.read('xl/styles.xml')

    ss_root = ET.fromstring(shared_strings_xml) if shared_strings_xml else None
    sheet_root = ET.fromstring(sheet_xml)

    # 3. 遍历 sheet1 cells,替换 inlineStr 或记录要改的 sharedString index
    index_to_translation = {}
    for row_el in sheet_root.iter(f'{_NS_PREFIX}row'):
        for c_el in row_el.findall(f'{_NS_PREFIX}c'):
            ref = c_el.get('r', '')
            cell_type = c_el.get('t', '')
            col_str = ''.join(c for c in ref if c.isalpha())
            row_num = int(''.join(c for c in ref if c.isdigit()))
            col_num = 0
            for ch in col_str:
                col_num = col_num * 26 + (ord(ch) - ord('A') + 1)
            key = (row_num - 1, col_num - 1)

            if cell_type == 's' and ss_root is not None:
                v_el = c_el.find(f'{_NS_PREFIX}v')
                if v_el is not None and v_el.text:
                    ss_index = int(v_el.text)
                    if key in translations and (row_num, col_num) not in skip_cells:
                        index_to_translation[ss_index] = translations[key]
            elif key in translations and (row_num, col_num) not in skip_cells:
                new_val = translations[key]
                if cell_type == 'inlineStr':
                    is_el = c_el.find(f'{_NS_PREFIX}is')
                    if is_el is not None:
                        for child in list(is_el):
                            is_el.remove(child)
                        t_el = ET.SubElement(is_el, f'{_NS_PREFIX}t')
                        t_el.text = new_val
                        t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
                else:
                    v_el = c_el.find(f'{_NS_PREFIX}v')
                    if v_el is not None:
                        c_el.remove(v_el)
                    c_el.set('t', 'inlineStr')
                    is_el = ET.SubElement(c_el, f'{_NS_PREFIX}is')
                    t_el = ET.SubElement(is_el, f'{_NS_PREFIX}t')
                    t_el.text = new_val
                    t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

    # 共享字符串表整段替换
    if ss_root is not None:
        for i, si in enumerate(ss_root.findall('ns:si', ns)):
            if i in index_to_translation:
                new_text = index_to_translation[i]
                for child in list(si):
                    si.remove(child)
                t_el = ET.SubElement(si, f'{_NS_PREFIX}t')
                t_el.text = new_text
                t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

    # 列宽:原宽 vs 翻译后估算宽度,取较大
    cols_el = sheet_root.find(f'{_NS_PREFIX}cols')
    original_col_widths = {}
    if cols_el is not None:
        for col_el in cols_el.findall(f'{_NS_PREFIX}col'):
            min_c = int(col_el.get('min', 0))
            max_c = int(col_el.get('max', 0))
            width = float(col_el.get('width', 8))
            for c in range(min_c, max_c + 1):
                original_col_widths[c] = width

    col_max_width = defaultdict(float)
    for (row, col), val in translations.items():
        r1, c1 = row + 1, col + 1
        if (r1, c1) in skip_cells:
            continue
        text_w = _estimate_text_width(val)
        span = merged_col_spans.get((r1, c1), 1)
        per_col_w = text_w / span
        for offset in range(span):
            col_max_width[c1 + offset] = max(col_max_width[c1 + offset], per_col_w)

    new_col_widths = dict(original_col_widths)
    for col, text_w in col_max_width.items():
        original_w = original_col_widths.get(col, 8)
        needed = text_w * 1.1 + 1
        if needed > original_w * 3:
            new_col_widths[col] = min(original_w * 1.5, 60)

    if cols_el is not None:
        sheet_root.remove(cols_el)
    if new_col_widths:
        cols_el = ET.SubElement(sheet_root, f'{_NS_PREFIX}cols')
        sheet_data = sheet_root.find(f'{_NS_PREFIX}sheetData')
        if sheet_data is not None:
            idx = list(sheet_root).index(sheet_data)
            sheet_root.remove(cols_el)
            sheet_root.insert(idx, cols_el)
        for col in sorted(new_col_widths.keys()):
            col_el = ET.SubElement(cols_el, f'{_NS_PREFIX}col')
            col_el.set('min', str(col))
            col_el.set('max', str(col))
            col_el.set('width', f'{new_col_widths[col]:.2f}')
            col_el.set('customWidth', '1')

    # 样式表:所有 cellXfs 加 wrapText(中文长行自动换行)
    styles_root = ET.fromstring(styles_xml)
    cell_xfs = styles_root.find(f'{_NS_PREFIX}cellXfs')
    if cell_xfs is not None:
        for xf in cell_xfs.findall(f'{_NS_PREFIX}xf'):
            alignment = xf.find(f'{_NS_PREFIX}alignment')
            if alignment is None:
                alignment = ET.SubElement(xf, f'{_NS_PREFIX}alignment')
            alignment.set('wrapText', '1')
            if alignment.get('shrinkToFit'):
                del alignment.attrib['shrinkToFit']
            xf.set('applyAlignment', '1')

    new_shared = ET.tostring(ss_root, encoding='unicode', xml_declaration=True) if ss_root is not None else None
    new_sheet = ET.tostring(sheet_root, encoding='unicode', xml_declaration=True)
    new_styles = ET.tostring(styles_root, encoding='unicode', xml_declaration=True)

    # 4. 重新打包 zip,改过的三项替换,其他全量复制
    with zipfile.ZipFile(original_path, 'r') as zin:
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                if item == 'xl/sharedStrings.xml' and new_shared is not None:
                    zout.writestr(item, new_shared)
                elif item == 'xl/worksheets/sheet1.xml':
                    zout.writestr(item, new_sheet)
                elif item == 'xl/styles.xml':
                    zout.writestr(item, new_styles)
                else:
                    zout.writestr(item, zin.read(item))
