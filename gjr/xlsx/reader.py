"""Excel 文件读取 + 可翻译性判断。

支持两种格式:
  - .xlsx  → openpyxl
  - .xls   → xlrd(老格式);写回时如果需要保留图片必须先转 .xlsx,见 ensure_xlsx()

LibreOffice 是**系统级依赖**(非 Python),用来把 .xls 转成 .xlsx 同时保留图片。
未装时 ensure_xlsx() 会抛异常,调用方需要降级到 xls_to_xlsx(会丢图片)。
"""
import re
import shutil
import subprocess
from pathlib import Path

import xlrd
from openpyxl import load_workbook


def _find_libreoffice_bin():
    """
    找本机可用的 LibreOffice 命令。

    跨平台命名:
      - Linux (apt):     libreoffice
      - macOS (brew):    soffice  (brew cask 只装这个名字,没有 libreoffice 别名)
      - Windows:         soffice.exe

    优先返回 PATH 里先找到的那个。全部找不到时返回 None,调用方负责抛出清晰错误。
    """
    for name in ("libreoffice", "soffice"):
        path = shutil.which(name)
        if path:
            return path
    # macOS brew cask 的备用硬编码路径(有时 PATH 没传到 subprocess)
    for fallback in (
        "/usr/local/bin/soffice",
        "/usr/local/bin/libreoffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/opt/homebrew/bin/soffice",
    ):
        if Path(fallback).exists():
            return fallback
    return None


def read_xls(file_path):
    """读 .xls(旧格式),返回 (cells_dict, nrows, ncols, sheet_name)。"""
    wb = xlrd.open_workbook(file_path)
    sheet = wb.sheet_by_index(0)
    cells = {}
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            val = sheet.cell_value(row, col)
            if val and str(val).strip():
                cells[(row, col)] = str(val).strip()
    return cells, sheet.nrows, sheet.ncols, sheet.name


def read_xlsx(file_path):
    """读 .xlsx,返回 (cells_dict, nrows, ncols, sheet_name)。"""
    wb = load_workbook(file_path)
    ws = wb.active
    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                cells[(cell.row - 1, cell.column - 1)] = str(cell.value).strip()
    wb.close()
    return cells, ws.max_row, ws.max_column, ws.title


def convert_xls_to_xlsx_libreoffice(xls_path):
    """
    用 LibreOffice headless 将 .xls 转 .xlsx,保留图片和格式。

    需要系统装 libreoffice:
      macOS:  brew install --cask libreoffice   (命令叫 soffice)
      Linux:  apt install libreoffice-calc      (命令叫 libreoffice)
    """
    bin_path = _find_libreoffice_bin()
    if bin_path is None:
        raise FileNotFoundError(
            "未找到 libreoffice/soffice 命令。"
            "macOS: brew install --cask libreoffice; "
            "Linux: apt install libreoffice-calc"
        )
    out_dir = str(Path(xls_path).parent)
    result = subprocess.run(
        [bin_path, "--headless", "--convert-to", "xlsx",
         str(xls_path), "--outdir", out_dir],
        capture_output=True, text=True, timeout=120,
    )
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice 转换失败: {result.stderr}")
    xlsx_path = str(Path(xls_path).with_suffix('.xlsx'))
    if not Path(xlsx_path).exists():
        raise RuntimeError("LibreOffice 转换后未找到 .xlsx 文件")
    return xlsx_path


def ensure_xlsx(file_path):
    """确保输入是 .xlsx;如果是 .xls,尝试用 libreoffice 转换(保留图片)。"""
    if Path(file_path).suffix.lower() != '.xls':
        return str(file_path)
    return convert_xls_to_xlsx_libreoffice(file_path)


def read_file(file_path):
    """按扩展名分派到对应读取函数。"""
    ext = Path(file_path).suffix.lower()
    if ext == '.xls':
        return read_xls(file_path)
    if ext == '.xlsx':
        return read_xlsx(file_path)
    raise ValueError(f"不支持的格式: {ext}")


_NUM_SYM_PAT = re.compile(r'^[\d./%±≤≥<>+\-]+$')
_PUNCT_PAT = re.compile(r'^[/\-—_=\.。，,]+$')


def is_translatable(text):
    """判断一个单元格内容是否值得送去翻译(纯数字/标点就跳过)。"""
    text = str(text).strip()
    if not text:
        return False
    if _NUM_SYM_PAT.match(text):
        return False
    if _PUNCT_PAT.match(text):
        return False
    return True
