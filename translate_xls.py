"""
XLS/XLSX 表格翻译工具

两阶段翻译：
  1. 提取术语表（LLM 看全局，确保术语一致）
  2. 分批翻译（按行分组，每批带完整表格结构作为上下文）

特点：
  - 发给 LLM 时附带 Markdown 表格结构，帮助理解行列关系
  - 只要求 LLM 输出 R/C 键值对，避免格式出错
  - 保留原表格格式、合并单元格、图片
  - Token 用量和费用统计

用法：
  python translate_xls.py input.xlsx -o output/
  python translate_xls.py input.xlsx -l en -o output/

依赖：openpyxl, xlrd, openai, python-dotenv
"""
import os
import re
import argparse
from pathlib import Path
from collections import defaultdict

import xlrd
from openpyxl import load_workbook, Workbook
from openai import OpenAI
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


def _get_api_key():
    """获取 API key：环境变量 → 配置文件 → 交互输入"""
    # 1. 环境变量
    key = os.getenv("OPENAI_API") or os.getenv("OPENAI_API_KEY")
    if key:
        return key

    # 2. 配置文件 ~/.translate_xls_key
    config_path = Path.home() / ".translate_xls_key"
    if config_path.exists():
        key = config_path.read_text().strip()
        if key:
            return key

    # 3. 交互输入并保存
    print("首次使用，请输入 OpenAI API Key:")
    key = input("API Key: ").strip()
    if not key:
        print("错误：API Key 不能为空")
        raise SystemExit(1)
    config_path.write_text(key)
    print(f"API Key 已保存到 {config_path}，下次无需再输入\n")
    return key


client = OpenAI(api_key=_get_api_key())

# 默认模型
DEFAULT_MODEL = "gpt-5.4-nano"    # 最便宜，试跑
# DEFAULT_MODEL = "gpt-5.4-mini"  # 性价比最优
# DEFAULT_MODEL = "gpt-5.4"      # 旗舰模型，质量最高

# 按语言指定模型
MODEL_PER_LANG = {
    "en": "gpt-5.4-nano",
    "th": "gpt-5.4-nano",
    "my": "gpt-5.4-nano",          # 如果缅甸语质量不够，改为 "gpt-5.4-mini"
}

# 价格 ($/1M tokens)
PRICING = {
    "gpt-5.4-nano": {"input": 0.20, "output": 1.25},
    "gpt-5.4-mini": {"input": 0.75, "output": 4.50},
    "gpt-5.4":      {"input": 2.50, "output": 15.00},
    "gpt-4o":       {"input": 2.50, "output": 10.00},
    "gpt-4o-mini":  {"input": 0.15, "output": 0.60},
}

# 支持的目标语言
LANGUAGES = {
    "en": "English",
    "th": "Thai (ภาษาไทย)",
    "my": "Burmese (မြန်မာဘာသာ)",
}

BATCH_SIZE = 25  # 每批翻译的单元格数（按行分组，实际可能略多或略少）

# Token 统计
_token_stats = {"input": 0, "output": 0, "calls": 0, "by_model": defaultdict(lambda: {"input": 0, "output": 0, "calls": 0})}


# ══════════════════════════════════════════════════════════════════════════════
# 文件读取
# ══════════════════════════════════════════════════════════════════════════════

def read_xls(file_path):
    """读取 .xls 文件，返回 {(row, col): value}"""
    wb = xlrd.open_workbook(file_path)
    sheet = wb.sheet_by_index(0)
    cells = {}
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            val = sheet.cell_value(row, col)
            if val and str(val).strip():
                cells[(row, col)] = str(val).strip()
    return cells, sheet.nrows, sheet.ncols, sheet.name


def read_xlsx(file_path):
    """读取 .xlsx 文件，返回 {(row, col): value}"""
    wb = load_workbook(file_path)
    ws = wb.active
    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                cells[(cell.row - 1, cell.column - 1)] = str(cell.value).strip()
    wb.close()
    return cells, ws.max_row, ws.max_column, ws.title


def read_file(file_path):
    """自动识别格式读取"""
    ext = Path(file_path).suffix.lower()
    if ext == '.xls':
        return read_xls(file_path)
    elif ext == '.xlsx':
        return read_xlsx(file_path)
    else:
        raise ValueError(f"不支持的格式: {ext}")


def is_translatable(text):
    """判断是否需要翻译（跳过纯数字、纯符号）"""
    text = str(text).strip()
    if not text:
        return False
    if re.match(r'^[\d./%±≤≥<>+\-]+$', text):
        return False
    if re.match(r'^[/\-—_=\.。，,]+$', text):
        return False
    return True


# ══════════════════════════════════════════════════════════════════════════════
# 表格结构生成
# ══════════════════════════════════════════════════════════════════════════════

def build_table_context(all_cells, nrows, ncols):
    """
    将全部单元格构建为 Markdown 表格文本，供 LLM 理解上下文结构。
    """
    # 找出实际使用的列
    used_cols = sorted(set(col for _, col in all_cells.keys()))
    if not used_cols:
        return ""

    lines = []
    for row in range(nrows):
        row_cells = []
        has_content = False
        for col in used_cols:
            val = all_cells.get((row, col), "")
            val = str(val).replace('\n', ' ').strip()[:30]  # 截断避免太长
            if val:
                has_content = True
            row_cells.append(val)
        if has_content:
            lines.append(f"| R{row+1} | " + " | ".join(row_cells) + " |")

    return '\n'.join(lines)


def format_cells_for_prompt(cells, keys=None):
    """将单元格格式化为 R/C 键值对"""
    if keys is None:
        keys = sorted(cells.keys())
    lines = []
    for row, col in keys:
        val = cells[(row, col)]
        val_escaped = val.replace('\n', '\\n')
        lines.append(f'R{row+1}C{col+1}: "{val_escaped}"')
    return '\n'.join(lines)


def group_by_rows(keys, max_cells=BATCH_SIZE):
    """
    按行分组，同一行的单元格不会被拆到不同批次。
    每批不超过 max_cells 个单元格。
    """
    row_groups = defaultdict(list)
    for row, col in keys:
        row_groups[row].append((row, col))

    batches = []
    current_batch = []
    for row in sorted(row_groups.keys()):
        row_keys = row_groups[row]
        if current_batch and len(current_batch) + len(row_keys) > max_cells:
            batches.append(current_batch)
            current_batch = []
        current_batch.extend(row_keys)
    if current_batch:
        batches.append(current_batch)

    return batches


# ══════════════════════════════════════════════════════════════════════════════
# LLM 调用
# ══════════════════════════════════════════════════════════════════════════════

def call_llm(system_prompt, user_prompt, temperature=0.3, lang=None):
    """调用 OpenAI API，记录 token 用量"""
    model = MODEL_PER_LANG.get(lang, DEFAULT_MODEL) if lang else DEFAULT_MODEL
    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        temperature=temperature,
    )

    # 统计 token
    usage = response.usage
    if usage:
        _token_stats["input"] += usage.prompt_tokens
        _token_stats["output"] += usage.completion_tokens
        _token_stats["calls"] += 1
        _token_stats["by_model"][model]["input"] += usage.prompt_tokens
        _token_stats["by_model"][model]["output"] += usage.completion_tokens
        _token_stats["by_model"][model]["calls"] += 1

    return response.choices[0].message.content


def print_token_stats():
    """打印 token 用量和费用统计"""
    print(f"\n{'='*50}")
    print("Token 用量和费用统计")
    print(f"{'='*50}")
    print(f"  总调用次数: {_token_stats['calls']}")
    print(f"  总 Input tokens:  {_token_stats['input']:,}")
    print(f"  总 Output tokens: {_token_stats['output']:,}")

    total_cost = 0
    for model, stats in _token_stats["by_model"].items():
        pricing = PRICING.get(model, {"input": 0, "output": 0})
        input_cost = stats["input"] / 1_000_000 * pricing["input"]
        output_cost = stats["output"] / 1_000_000 * pricing["output"]
        model_cost = input_cost + output_cost
        total_cost += model_cost
        print(f"\n  模型: {model} ({stats['calls']} 次调用)")
        print(f"    Input:  {stats['input']:>8,} tokens × ${pricing['input']}/1M = ${input_cost:.4f}")
        print(f"    Output: {stats['output']:>8,} tokens × ${pricing['output']}/1M = ${output_cost:.4f}")
        print(f"    小计: ${model_cost:.4f}")

    print(f"\n  总费用: ${total_cost:.4f}")


# ══════════════════════════════════════════════════════════════════════════════
# 翻译逻辑
# ══════════════════════════════════════════════════════════════════════════════

def extract_glossary(cells, target_lang, output_dir=None):
    """
    阶段一：提取术语对照表（每种语言单独提取）。
    """
    all_text = format_cells_for_prompt(cells)
    lang_name = LANGUAGES[target_lang]

    system_prompt = f"你是工业制造领域的专业翻译专家，精通中文和{lang_name}。"

    # 泰语/缅甸语的术语表增加 fallback 指导
    fallback_glossary_rule = ""
    if target_lang in ("th", "my"):
        fallback_glossary_rule = f"""
- 对于难以直接翻译为{lang_name}的术语，请按优先级尝试：
  1. 分词后逐部分翻译为{lang_name}（如"底钢侧板"→拆分为"底部"+"钢"+"侧板"分别翻译）
  2. 无法用{lang_name}表达的部分用英文
  3. 中国人名/地名用{lang_name}音译或拼音
  4. 只有完全无法处理时才保留中文
  在注意事项中标注你选择了哪种 fallback 方式及原因"""

    user_prompt = f"""以下是一份工业制造作业指导书的表格内容。
请提取所有专业术语和重复出现的关键词（约20-40个），
给出中文 → {lang_name} 的对照翻译。

要求：
- 只提取专业术语和关键词，不要翻译整个句子
- 术语翻译要准确、专业{fallback_glossary_rule}
- 对于在{lang_name}中容易混淆、有多种译法、或需要特别注意的术语，
  请在后面用括号标注翻译难点和选择该译法的理由

输出格式（每行一个术语）：
中文 | {lang_name}翻译 | 注意事项（如有）

示例：
毛刺 | Burr | (钣金行业专用，不要译为 flash/fin)
首检 | First Article Inspection | (FAI，行业标准缩写)

---表格内容---
{all_text}"""

    print(f"  提取术语表 ({lang_name})...")
    result = call_llm(system_prompt, user_prompt, temperature=0.1, lang=target_lang)
    print(f"  术语表已生成 ({len(result.splitlines())} 行)")

    # 保存术语表
    if output_dir:
        glossary_path = Path(output_dir) / f"glossary_{target_lang}.txt"
        glossary_path.write_text(result, encoding='utf-8')
        print(f"  术语表已保存: {glossary_path}")

    return result


def translate_batch(batch_keys, all_cells, table_context, glossary, target_lang):
    """
    阶段二：翻译一批单元格。

    发送给 LLM 的内容：
      1. 完整表格结构（Markdown 格式，仅供理解上下文）
      2. 需要翻译的 R/C 键值对（这些才是要翻译的）
    """
    lang_name = LANGUAGES[target_lang]
    batch_text = format_cells_for_prompt(all_cells, keys=batch_keys)

    # 泰语和缅甸语添加专有名词 fallback 策略
    fallback_rule = ""
    if target_lang in ("th", "my"):
        fallback_rule = f"""
- 对于难以直接翻译的专有名词（如公司名、产品名、技术术语），按以下优先级处理：
  1. 先尝试分词：将词拆分成更小的部分，逐部分翻译为{lang_name}
  2. 如果分词后仍无法用{lang_name}表达，则翻译为英文
  3. 如果英文也不合适（如中国人名），则使用拼音
  4. 以上都不行，才保留中文原文
  例如："海尔冰箱" → 分词为 "海尔"+"冰箱" → 海尔音译 + 冰箱翻译为{lang_name}
  例如："PCM板" → PCM 保留 + "板"翻译为{lang_name}"""

    system_prompt = f"""你是工业制造领域的专业翻译专家，精通中文和{lang_name}。
将中文翻译为 {lang_name}。你必须尽最大努力输出{lang_name}，避免输出其他语言的文字。

必须遵守以下术语表（保持术语一致性）：
{glossary}

翻译规则：
- 保持原文的换行符(\\n)
- 纯数字、编号、型号不翻译，原样保留
- 翻译要自然流畅，符合{lang_name}的表达习惯
- 严格按照术语表中的翻译{fallback_rule}"""

    user_prompt = f"""以下是完整表格的结构（仅供你理解上下文和行列关系，不要翻译这部分）：

{table_context}

---

请翻译以下单元格为 {lang_name}。
输出格式：每行 R{{行}}C{{列}}: "译文"
只输出翻译结果，不要输出其他内容。

{batch_text}"""

    result = call_llm(system_prompt, user_prompt, temperature=0.2, lang=target_lang)
    return result


def parse_translation_result(result_text):
    """解析 LLM 返回的翻译结果"""
    translations = {}
    for line in result_text.strip().splitlines():
        line = line.strip()
        match = re.match(r'R(\d+)C(\d+):\s*"(.+)"', line)
        if not match:
            match = re.match(r'R(\d+)C(\d+):\s*(.+)', line)
        if match:
            row = int(match.group(1)) - 1
            col = int(match.group(2)) - 1
            val = match.group(3).strip().strip('"')
            val = val.replace('\\n', '\n')
            translations[(row, col)] = val
    return translations


# ══════════════════════════════════════════════════════════════════════════════
# 文件保存
# ══════════════════════════════════════════════════════════════════════════════

def xls_to_xlsx(xls_path):
    """将 .xls 转为 .xlsx（尽量保留合并单元格，但图片会丢失）"""
    wb_old = xlrd.open_workbook(xls_path, formatting_info=True)
    sheet = wb_old.sheet_by_index(0)

    wb_new = Workbook()
    ws = wb_new.active
    ws.title = sheet.name

    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            val = sheet.cell_value(row, col)
            if val:
                ws.cell(row=row+1, column=col+1, value=val)

    for rlo, rhi, clo, chi in sheet.merged_cells:
        ws.merge_cells(
            start_row=rlo+1, start_column=clo+1,
            end_row=rhi, end_column=chi
        )

    return wb_new


_LANG_FONT = {
    "th": "TH Sarabun New",
    "my": "Myanmar Text",
}


def save_translated(original_path, translations, output_path, target_lang=None):
    """保存翻译结果，保留原文件的格式和图片。"""
    ext = Path(original_path).suffix.lower()

    if ext == '.xlsx':
        _save_xlsx_with_images(original_path, translations, output_path, target_lang)
    else:
        wb = xls_to_xlsx(original_path)
        ws = wb.active
        for (row, col), val in translations.items():
            ws.cell(row=row+1, column=col+1, value=val)
        wb.save(output_path)
        print(f"  已保存: {output_path}")


def _estimate_text_width(text):
    """估算文本在 Excel 中的显示宽度（字符数）。中文字符算 2，其他算 1。"""
    width = 0
    for ch in str(text):
        if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f':
            width += 2  # 中日韩字符
        elif '\u0e00' <= ch <= '\u0e7f':
            width += 1.5  # 泰语
        elif '\u1000' <= ch <= '\u109f':
            width += 1.5  # 缅甸语
        else:
            width += 1
    return width


def _save_xlsx_with_images(original_path, translations, output_path, target_lang=None):
    """
    直接操作 xlsx zip 包，替换文字，保留图片/绘图/样式。
    同时自动调整列宽、行高。
    """
    import zipfile
    import xml.etree.ElementTree as ET

    wb = load_workbook(original_path)
    ws = wb.active

    # 找出合并单元格信息
    skip_cells = set()
    merged_col_spans = {}  # (min_row, min_col) → 合并的列数
    for merged_range in ws.merged_cells.ranges:
        col_span = merged_range.max_col - merged_range.min_col + 1
        merged_col_spans[(merged_range.min_row, merged_range.min_col)] = col_span
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if (row, col) != (merged_range.min_row, merged_range.min_col):
                    skip_cells.add((row, col))
    wb.close()

    NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    ns = {'ns': NS}
    ET.register_namespace('', NS)

    with zipfile.ZipFile(original_path, 'r') as zin:
        has_shared_strings = 'xl/sharedStrings.xml' in zin.namelist()
        shared_strings_xml = zin.read('xl/sharedStrings.xml') if has_shared_strings else None
        sheet_xml = zin.read('xl/worksheets/sheet1.xml')
        styles_xml = zin.read('xl/styles.xml')

    # ── 解析 sheet ──
    ss_root = ET.fromstring(shared_strings_xml) if shared_strings_xml else None
    sheet_root = ET.fromstring(sheet_xml)

    # 建立 cell ref → (row, col) 的映射，替换翻译
    cell_ref_to_pos = {}
    index_to_translation = {}
    for row_el in sheet_root.iter(f'{{{NS}}}row'):
        for c_el in row_el.findall(f'{{{NS}}}c'):
            ref = c_el.get('r', '')
            cell_type = c_el.get('t', '')
            col_str = ''.join(c for c in ref if c.isalpha())
            row_num = int(''.join(c for c in ref if c.isdigit()))
            col_num = 0
            for ch in col_str:
                col_num = col_num * 26 + (ord(ch) - ord('A') + 1)
            cell_ref_to_pos[ref] = (row_num, col_num)
            key = (row_num - 1, col_num - 1)

            if cell_type == 's' and ss_root is not None:
                # sharedString 引用
                v_el = c_el.find(f'{{{NS}}}v')
                if v_el is not None and v_el.text:
                    ss_index = int(v_el.text)
                    if key in translations and (row_num, col_num) not in skip_cells:
                        index_to_translation[ss_index] = translations[key]
            elif key in translations and (row_num, col_num) not in skip_cells:
                # 内联值或普通值 → 直接在 sheet XML 中替换
                new_val = translations[key]
                if cell_type == 'inlineStr':
                    is_el = c_el.find(f'{{{NS}}}is')
                    if is_el is not None:
                        for child in list(is_el):
                            is_el.remove(child)
                        t_el = ET.SubElement(is_el, f'{{{NS}}}t')
                        t_el.text = new_val
                        t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
                else:
                    # 普通 <v> 值 → 转为 inlineStr
                    v_el = c_el.find(f'{{{NS}}}v')
                    if v_el is not None:
                        c_el.remove(v_el)
                    c_el.set('t', 'inlineStr')
                    is_el = ET.SubElement(c_el, f'{{{NS}}}is')
                    t_el = ET.SubElement(is_el, f'{{{NS}}}t')
                    t_el.text = new_val
                    t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

    if ss_root is not None:
        for i, si in enumerate(ss_root.findall('ns:si', ns)):
            if i in index_to_translation:
                new_text = index_to_translation[i]
                for child in list(si):
                    si.remove(child)
                t_el = ET.SubElement(si, f'{{{NS}}}t')
                t_el.text = new_text
                t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

    # ── 列宽策略：保持原始布局，优先用 wrapText + shrinkToFit 适配 ──

    # 读取原始列宽
    cols_el = sheet_root.find(f'{{{NS}}}cols')
    original_col_widths = {}
    if cols_el is not None:
        for col_el in cols_el.findall(f'{{{NS}}}col'):
            min_c = int(col_el.get('min', 0))
            max_c = int(col_el.get('max', 0))
            width = float(col_el.get('width', 8))
            for c in range(min_c, max_c + 1):
                original_col_widths[c] = width

    # 计算每列翻译后的最大文本宽度
    col_max_width = defaultdict(float)
    for (row, col), val in translations.items():
        r1, c1 = row + 1, col + 1
        if (r1, c1) in skip_cells:
            continue
        text_w = _estimate_text_width(val)
        span = merged_col_spans.get((r1, c1), 1)
        per_col_w = text_w / span
        for offset in range(span):
            col_max_width[c1 + offset] = max(col_max_width[c1 + offset], per_col_w)

    # 只在译文比原文长 3 倍以上时才扩列宽（极端情况兜底）
    new_col_widths = dict(original_col_widths)
    for col, text_w in col_max_width.items():
        original_w = original_col_widths.get(col, 8)
        needed = text_w * 1.1 + 1
        if needed > original_w * 3:
            new_col_widths[col] = min(original_w * 1.5, 60)

    # 写回 <cols>（保持原始列宽为主）
    if cols_el is not None:
        sheet_root.remove(cols_el)
    if new_col_widths:
        cols_el = ET.SubElement(sheet_root, f'{{{NS}}}cols')
        sheet_data = sheet_root.find(f'{{{NS}}}sheetData')
        if sheet_data is not None:
            idx = list(sheet_root).index(sheet_data)
            sheet_root.remove(cols_el)
            sheet_root.insert(idx, cols_el)

        for col in sorted(new_col_widths.keys()):
            col_el = ET.SubElement(cols_el, f'{{{NS}}}col')
            col_el.set('min', str(col))
            col_el.set('max', str(col))
            col_el.set('width', f'{new_col_widths[col]:.2f}')
            col_el.set('customWidth', '1')

    # ── 修改样式：启用 wrapText 自动换行 ──
    styles_root = ET.fromstring(styles_xml)

    cell_xfs = styles_root.find(f'{{{NS}}}cellXfs')
    if cell_xfs is not None:
        for xf in cell_xfs.findall(f'{{{NS}}}xf'):
            alignment = xf.find(f'{{{NS}}}alignment')
            if alignment is None:
                alignment = ET.SubElement(xf, f'{{{NS}}}alignment')
            alignment.set('wrapText', '1')
            if alignment.get('shrinkToFit'):
                del alignment.attrib['shrinkToFit']
            xf.set('applyAlignment', '1')

    # ── 替换字体：中文字体不含泰语/缅甸语字形，会导致字符间距异常 ──
    target_font = _LANG_FONT.get(target_lang)
    if target_font:
        fonts_el = styles_root.find(f'{{{NS}}}fonts')
        if fonts_el is not None:
            for font in fonts_el.findall(f'{{{NS}}}font'):
                name_el = font.find(f'{{{NS}}}name')
                if name_el is not None:
                    name_el.set('val', target_font)

    # ── 写入 zip ──
    new_shared = ET.tostring(ss_root, encoding='unicode', xml_declaration=True) if ss_root is not None else None
    new_sheet = ET.tostring(sheet_root, encoding='unicode', xml_declaration=True)
    new_styles = ET.tostring(styles_root, encoding='unicode', xml_declaration=True)

    with zipfile.ZipFile(original_path, 'r') as zin:
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                if item == 'xl/sharedStrings.xml' and new_shared is not None:
                    zout.writestr(item, new_shared)
                elif item == 'xl/worksheets/sheet1.xml':
                    zout.writestr(item, new_sheet)
                elif item == 'xl/styles.xml':
                    zout.writestr(item, new_styles)
                else:
                    zout.writestr(item, zin.read(item))

    print(f"  已保存: {output_path} (图片已保留, 列宽已调整)")


# ══════════════════════════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════════════════════════

def translate_file(input_path, target_langs=None, output_dir=None):
    """主函数：翻译 XLS/XLSX 文件。"""
    from datetime import datetime

    if target_langs is None:
        target_langs = ['en', 'th', 'my']

    stem = Path(input_path).stem
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # 在 output_dir 下创建 {文件名}_{时间戳} 子目录
    base_dir = Path(output_dir) if output_dir else Path(input_path).parent
    output_dir = str(base_dir / f"{stem}_{timestamp}")

    os.makedirs(output_dir, exist_ok=True)
    print(f"输出目录: {output_dir}")

    ext = Path(input_path).suffix.lower()

    if ext == '.xls':
        print("  提示: .xls 格式无法保留图片，建议先用 WPS/Excel 另存为 .xlsx")

    # 读取文件
    print(f"读取: {input_path}")
    cells, nrows, ncols, sheet_name = read_file(input_path)

    # 过滤需要翻译的单元格
    translatable = {k: v for k, v in cells.items() if is_translatable(v)}
    print(f"  共 {len(cells)} 个单元格，{len(translatable)} 个需要翻译")
    print(f"  默认模型: {DEFAULT_MODEL}，批次大小: ~{BATCH_SIZE}")

    # 构建表格上下文（Markdown 格式，所有批次共用）
    table_context = build_table_context(cells, nrows, ncols)

    # 分语言处理
    for lang in target_langs:
        lang_name = LANGUAGES[lang]
        print(f"\n{'='*50}")
        print(f"翻译为 {lang_name} (模型: {MODEL_PER_LANG.get(lang, DEFAULT_MODEL)})")

        # 阶段一：提取该语言的术语表
        glossary = extract_glossary(translatable, lang, output_dir)

        # 按行分组（同一行的单元格不会被拆分）
        keys = sorted(translatable.keys())
        batches = group_by_rows(keys, BATCH_SIZE)

        all_translations = {}

        for batch_idx, batch_keys in enumerate(batches):
            # 显示本批次包含的行范围
            rows_in_batch = sorted(set(r for r, c in batch_keys))
            row_range = f"R{rows_in_batch[0]+1}-R{rows_in_batch[-1]+1}"
            print(f"  批次 {batch_idx+1}/{len(batches)} ({len(batch_keys)} 个单元格, {row_range})...", end='', flush=True)
            try:
                result = translate_batch(batch_keys, translatable, table_context, glossary, lang)
                parsed = parse_translation_result(result)
                all_translations.update(parsed)
                print(f" 完成 ({len(parsed)} 个)")
            except Exception as e:
                print(f" 失败: {e}")
                try:
                    print(f"  重试...", end='', flush=True)
                    result = translate_batch(batch_keys, translatable, table_context, glossary, lang)
                    parsed = parse_translation_result(result)
                    all_translations.update(parsed)
                    print(f" 完成 ({len(parsed)} 个)")
                except Exception as e2:
                    print(f" 再次失败: {e2}，跳过此批次")

        # 补上不需要翻译的单元格
        for k, v in cells.items():
            if k not in all_translations:
                all_translations[k] = v

        # 保存
        output_path = Path(output_dir) / f"{stem}_{lang}.xlsx"
        save_translated(input_path, all_translations, str(output_path), target_lang=lang)

        print(f"  {lang_name}: {len(all_translations)} 个单元格已翻译")

    # 打印统计
    print_token_stats()


def convert_xls_to_xlsx(xls_path):
    """将 .xls 转为 .xlsx（保留合并单元格和数据，图片可能丢失）"""
    xlsx_path = str(Path(xls_path).with_suffix('.xlsx'))
    wb = xls_to_xlsx(xls_path)
    wb.save(xlsx_path)
    print(f"  已转换: {xls_path} → {xlsx_path}")
    return xlsx_path


def ensure_xlsx(file_path):
    """确保输入是 xlsx 格式，如果是 xls 自动转换"""
    ext = Path(file_path).suffix.lower()
    if ext == '.xls':
        print(f"  检测到 .xls 格式，自动转换为 .xlsx...")
        return convert_xls_to_xlsx(file_path)
    return file_path


def show_api_key_info():
    """显示当前 API Key 信息"""
    config_path = Path.home() / ".translate_xls_key"
    key = os.getenv("OPENAI_API") or os.getenv("OPENAI_API_KEY")
    source = "环境变量"
    if not key and config_path.exists():
        key = config_path.read_text().strip()
        source = f"配置文件 ({config_path})"
    if key:
        masked = key[:8] + "..." + key[-4:]
        print(f"  当前 API Key: {masked} (来源: {source})")
    else:
        print("  当前未设置 API Key")


def set_api_key():
    """设置新的 API Key"""
    config_path = Path.home() / ".translate_xls_key"
    show_api_key_info()
    print()
    new_key = input("  请输入新的 API Key (直接回车取消): ").strip()
    if new_key:
        config_path.write_text(new_key)
        print(f"  API Key 已保存到 {config_path}")
        # 重新初始化 client
        global client
        client = OpenAI(api_key=new_key)
    else:
        print("  已取消")


def interactive_translate():
    """交互式翻译"""
    file_path = input("  请输入文件路径 (支持拖入文件): ").strip().strip("'\"")

    if not file_path:
        print("  已取消")
        return

    if not os.path.exists(file_path):
        print(f"  错误: 文件不存在 - {file_path}")
        return

    # 自动转换 xls → xlsx
    file_path = ensure_xlsx(file_path)

    # 输出目录默认为文件所在目录
    output_dir = str(Path(file_path).parent)

    print(f"  文件: {Path(file_path).name}")
    print(f"  输出目录: {output_dir}")
    print(f"  目标语言: English, Thai, Burmese")
    print()

    confirm = input("  确认开始翻译? (Y/n): ").strip().lower()
    if confirm in ('n', 'no'):
        print("  已取消")
        return

    translate_file(file_path, ['en', 'th', 'my'], output_dir)


def main_menu():
    """主菜单"""
    print()
    print("=" * 50)
    print("  XLS/XLSX 表格翻译工具")
    print("  英语 / 泰语 / 缅甸语")
    print("=" * 50)

    while True:
        print()
        print("  1. 翻译文件 (英语/泰语/缅甸语)")
        print("  2. 设置 API Key")
        print("  3. 退出")
        print()

        choice = input("  请选择 (1/2/3): ").strip()

        if choice == '1':
            try:
                interactive_translate()
            except KeyboardInterrupt:
                print("\n  已中断")
            except Exception as e:
                print(f"\n  翻译出错: {e}")
        elif choice == '2':
            set_api_key()
        elif choice == '3':
            print("  再见!")
            break
        else:
            print("  无效选择，请输入 1、2 或 3")


if __name__ == '__main__':
    # 如果有命令行参数，走参数模式；否则走交互式菜单
    if len(os.sys.argv) > 1:
        parser = argparse.ArgumentParser(description='XLS/XLSX 表格翻译工具')
        parser.add_argument('input', help='输入文件路径 (.xls 或 .xlsx)')
        parser.add_argument('-l', '--langs', nargs='+', default=['en', 'th', 'my'],
                            choices=list(LANGUAGES.keys()),
                            help='目标语言 (默认: en th my)')
        parser.add_argument('-o', '--output', help='输出目录 (默认: 输入文件所在目录)')
        args = parser.parse_args()
        file_path = ensure_xlsx(args.input)
        translate_file(file_path, args.langs, args.output)
    else:
        main_menu()
