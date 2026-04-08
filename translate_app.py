"""
XLS/XLSX 表格翻译工具 - Streamlit Web 版

基于 translate_xls.py 的核心逻辑，提供网页界面。
用法：streamlit run translate_app.py
"""
import os
import re
import tempfile
import zipfile
import subprocess
import xml.etree.ElementTree as ET
from pathlib import Path
from collections import defaultdict
from datetime import datetime

import streamlit as st
import xlrd
from openpyxl import load_workbook, Workbook
from openai import OpenAI

# ══════════════════════════════════════════════════════════════════════════════
# 配置
# ══════════════════════════════════════════════════════════════════════════════

DEFAULT_MODEL = "gpt-5.4-nano"

MODEL_PER_LANG = {
    "en": "gpt-5.4-nano",
    "th": "gpt-5.4-nano",
    "my": "gpt-5.4-nano",
}

PRICING = {
    "gpt-5.4-nano": {"input": 0.20, "output": 1.25},
    "gpt-5.4-mini": {"input": 0.75, "output": 4.50},
    "gpt-5.4":      {"input": 2.50, "output": 15.00},
    "gpt-4o":       {"input": 2.50, "output": 10.00},
    "gpt-4o-mini":  {"input": 0.15, "output": 0.60},
}

LANGUAGES = {
    "en": "English",
    "th": "Thai (ภาษาไทย)",
    "my": "Burmese (မြန်မာဘာသာ)",
}

BATCH_SIZE = 25


# ══════════════════════════════════════════════════════════════════════════════
# 文件读取（复用 translate_xls.py 逻辑）
# ══════════════════════════════════════════════════════════════════════════════

def read_xls(file_path):
    wb = xlrd.open_workbook(file_path)
    sheet = wb.sheet_by_index(0)
    cells = {}
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            val = sheet.cell_value(row, col)
            if val and str(val).strip():
                cells[(row, col)] = str(val).strip()
    return cells, sheet.nrows, sheet.ncols, sheet.name


def read_xlsx(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                cells[(cell.row - 1, cell.column - 1)] = str(cell.value).strip()
    wb.close()
    return cells, ws.max_row, ws.max_column, ws.title


def convert_xls_to_xlsx_libreoffice(xls_path):
    """用 LibreOffice 将 .xls 转为 .xlsx，保留图片和格式"""
    out_dir = str(Path(xls_path).parent)
    result = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "xlsx", xls_path, "--outdir", out_dir],
        capture_output=True, text=True, timeout=60,
    )
    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice 转换失败: {result.stderr}")
    xlsx_path = str(Path(xls_path).with_suffix('.xlsx'))
    if not Path(xlsx_path).exists():
        raise RuntimeError("LibreOffice 转换后未找到 .xlsx 文件")
    return xlsx_path


def ensure_xlsx(file_path):
    """确保输入是 xlsx 格式。xls 自动用 LibreOffice 转换（保留图片）"""
    if Path(file_path).suffix.lower() != '.xls':
        return file_path
    return convert_xls_to_xlsx_libreoffice(file_path)


def read_file(file_path):
    ext = Path(file_path).suffix.lower()
    if ext == '.xls':
        return read_xls(file_path)
    elif ext == '.xlsx':
        return read_xlsx(file_path)
    else:
        raise ValueError(f"不支持的格式: {ext}")


def is_translatable(text):
    text = str(text).strip()
    if not text:
        return False
    if re.match(r'^[\d./%±≤≥<>+\-]+$', text):
        return False
    if re.match(r'^[/\-—_=\.。，,]+$', text):
        return False
    return True


# ══════════════════════════════════════════════════════════════════════════════
# 表格结构 & LLM
# ══════════════════════════════════════════════════════════════════════════════

def build_table_context(all_cells, nrows, ncols):
    used_cols = sorted(set(col for _, col in all_cells.keys()))
    if not used_cols:
        return ""
    lines = []
    for row in range(nrows):
        row_cells = []
        has_content = False
        for col in used_cols:
            val = all_cells.get((row, col), "")
            val = str(val).replace('\n', ' ').strip()[:30]
            if val:
                has_content = True
            row_cells.append(val)
        if has_content:
            lines.append(f"| R{row+1} | " + " | ".join(row_cells) + " |")
    return '\n'.join(lines)


def format_cells_for_prompt(cells, keys=None):
    if keys is None:
        keys = sorted(cells.keys())
    lines = []
    for row, col in keys:
        val = cells[(row, col)]
        val_escaped = val.replace('\n', '\\n')
        lines.append(f'R{row+1}C{col+1}: "{val_escaped}"')
    return '\n'.join(lines)


def group_by_rows(keys, max_cells=BATCH_SIZE):
    row_groups = defaultdict(list)
    for row, col in keys:
        row_groups[row].append((row, col))
    batches = []
    current_batch = []
    for row in sorted(row_groups.keys()):
        row_keys = row_groups[row]
        if current_batch and len(current_batch) + len(row_keys) > max_cells:
            batches.append(current_batch)
            current_batch = []
        current_batch.extend(row_keys)
    if current_batch:
        batches.append(current_batch)
    return batches


def call_llm(client, system_prompt, user_prompt, temperature=0.3, lang=None):
    model = MODEL_PER_LANG.get(lang, DEFAULT_MODEL) if lang else DEFAULT_MODEL
    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        temperature=temperature,
    )
    usage = response.usage
    tokens = {"input": 0, "output": 0, "model": model}
    if usage:
        tokens["input"] = usage.prompt_tokens
        tokens["output"] = usage.completion_tokens
    return response.choices[0].message.content, tokens


# ══════════════════════════════════════════════════════════════════════════════
# 翻译逻辑
# ══════════════════════════════════════════════════════════════════════════════

def extract_glossary(client, cells, target_lang):
    all_text = format_cells_for_prompt(cells)
    lang_name = LANGUAGES[target_lang]

    system_prompt = f"你是工业制造领域的专业翻译专家，精通中文和{lang_name}。"

    fallback_glossary_rule = ""
    if target_lang in ("th", "my"):
        fallback_glossary_rule = f"""
- 对于难以直接翻译为{lang_name}的术语，请按优先级尝试：
  1. 分词后逐部分翻译为{lang_name}（如"底钢侧板"→拆分为"底部"+"钢"+"侧板"分别翻译）
  2. 无法用{lang_name}表达的部分用英文
  3. 中国人名/地名用{lang_name}音译或拼音
  4. 只有完全无法处理时才保留中文
  在注意事项中标注你选择了哪种 fallback 方式及原因"""

    user_prompt = f"""以下是一份工业制造作业指导书的表格内容。
请提取所有专业术语和重复出现的关键词（约20-40个），
给出中文 → {lang_name} 的对照翻译。

要求：
- 只提取专业术语和关键词，不要翻译整个句子
- 术语翻译要准确、专业{fallback_glossary_rule}
- 对于在{lang_name}中容易混淆、有多种译法、或需要特别注意的术语，
  请在后面用括号标注翻译难点和选择该译法的理由

输出格式（每行一个术语）：
中文 | {lang_name}翻译 | 注意事项（如有）

示例：
毛刺 | Burr | (钣金行业专用，不要译为 flash/fin)
首检 | First Article Inspection | (FAI，行业标准缩写)

---表格内容---
{all_text}"""

    result, tokens = call_llm(client, system_prompt, user_prompt, temperature=0.1, lang=target_lang)
    return result, tokens


def translate_batch(client, batch_keys, all_cells, table_context, glossary, target_lang):
    lang_name = LANGUAGES[target_lang]
    batch_text = format_cells_for_prompt(all_cells, keys=batch_keys)

    fallback_rule = ""
    if target_lang in ("th", "my"):
        fallback_rule = f"""
- 对于难以直接翻译的专有名词（如公司名、产品名、技术术语），按以下优先级处理：
  1. 先尝试分词：将词拆分成更小的部分，逐部分翻译为{lang_name}
  2. 如果分词后仍无法用{lang_name}表达，则翻译为英文
  3. 如果英文也不合适（如中国人名），则使用拼音
  4. 以上都不行，才保留中文原文
  例如："海尔冰箱" → 分词为 "海尔"+"冰箱" → 海尔音译 + 冰箱翻译为{lang_name}
  例如："PCM板" → PCM 保留 + "板"翻译为{lang_name}"""

    system_prompt = f"""你是工业制造领域的专业翻译专家，精通中文和{lang_name}。
将中文翻译为 {lang_name}。你必须尽最大努力输出{lang_name}，避免输出其他语言的文字。

必须遵守以下术语表（保持术语一致性）：
{glossary}

翻译规则：
- 保持原文的换行符(\\n)
- 纯数字、编号、型号不翻译，原样保留
- 翻译要自然流畅，符合{lang_name}的表达习惯
- 严格按照术语表中的翻译{fallback_rule}"""

    user_prompt = f"""以下是完整表格的结构（仅供你理解上下文和行列关系，不要翻译这部分）：

{table_context}

---

请翻译以下单元格为 {lang_name}。
输出格式：每行 R{{行}}C{{列}}: "译文"
只输出翻译结果，不要输出其他内容。

{batch_text}"""

    result, tokens = call_llm(client, system_prompt, user_prompt, temperature=0.2, lang=target_lang)
    return result, tokens


def parse_translation_result(result_text):
    translations = {}
    for line in result_text.strip().splitlines():
        line = line.strip()
        match = re.match(r'R(\d+)C(\d+):\s*"(.+)"', line)
        if not match:
            match = re.match(r'R(\d+)C(\d+):\s*(.+)', line)
        if match:
            row = int(match.group(1)) - 1
            col = int(match.group(2)) - 1
            val = match.group(3).strip().strip('"')
            val = val.replace('\\n', '\n')
            translations[(row, col)] = val
    return translations


# ══════════════════════════════════════════════════════════════════════════════
# 文件保存
# ══════════════════════════════════════════════════════════════════════════════

def xls_to_xlsx(xls_path):
    wb_old = xlrd.open_workbook(xls_path, formatting_info=True)
    sheet = wb_old.sheet_by_index(0)
    wb_new = Workbook()
    ws = wb_new.active
    ws.title = sheet.name
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            val = sheet.cell_value(row, col)
            if val:
                ws.cell(row=row+1, column=col+1, value=val)
    for rlo, rhi, clo, chi in sheet.merged_cells:
        ws.merge_cells(start_row=rlo+1, start_column=clo+1, end_row=rhi, end_column=chi)
    return wb_new


def _estimate_text_width(text):
    width = 0
    for ch in str(text):
        if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f':
            width += 2
        elif '\u0e00' <= ch <= '\u0e7f':
            width += 1.5
        elif '\u1000' <= ch <= '\u109f':
            width += 1.5
        else:
            width += 1
    return width


def save_translated_xlsx(original_path, translations, output_path):
    ext = Path(original_path).suffix.lower()
    if ext == '.xlsx':
        _save_xlsx_with_images(original_path, translations, output_path)
    else:
        wb = xls_to_xlsx(original_path)
        ws = wb.active
        for (row, col), val in translations.items():
            ws.cell(row=row+1, column=col+1, value=val)
        wb.save(output_path)


def _save_xlsx_with_images(original_path, translations, output_path):
    wb = load_workbook(original_path)
    ws = wb.active

    skip_cells = set()
    merged_col_spans = {}
    for merged_range in ws.merged_cells.ranges:
        col_span = merged_range.max_col - merged_range.min_col + 1
        merged_col_spans[(merged_range.min_row, merged_range.min_col)] = col_span
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if (row, col) != (merged_range.min_row, merged_range.min_col):
                    skip_cells.add((row, col))
    wb.close()

    NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    ns = {'ns': NS}
    ET.register_namespace('', NS)

    with zipfile.ZipFile(original_path, 'r') as zin:
        has_shared_strings = 'xl/sharedStrings.xml' in zin.namelist()
        shared_strings_xml = zin.read('xl/sharedStrings.xml') if has_shared_strings else None
        sheet_xml = zin.read('xl/worksheets/sheet1.xml')
        styles_xml = zin.read('xl/styles.xml')

    ss_root = ET.fromstring(shared_strings_xml) if shared_strings_xml else None
    sheet_root = ET.fromstring(sheet_xml)

    index_to_translation = {}
    for row_el in sheet_root.iter(f'{{{NS}}}row'):
        for c_el in row_el.findall(f'{{{NS}}}c'):
            ref = c_el.get('r', '')
            cell_type = c_el.get('t', '')
            col_str = ''.join(c for c in ref if c.isalpha())
            row_num = int(''.join(c for c in ref if c.isdigit()))
            col_num = 0
            for ch in col_str:
                col_num = col_num * 26 + (ord(ch) - ord('A') + 1)
            key = (row_num - 1, col_num - 1)

            if cell_type == 's' and ss_root is not None:
                v_el = c_el.find(f'{{{NS}}}v')
                if v_el is not None and v_el.text:
                    ss_index = int(v_el.text)
                    if key in translations and (row_num, col_num) not in skip_cells:
                        index_to_translation[ss_index] = translations[key]
            elif key in translations and (row_num, col_num) not in skip_cells:
                new_val = translations[key]
                if cell_type == 'inlineStr':
                    is_el = c_el.find(f'{{{NS}}}is')
                    if is_el is not None:
                        for child in list(is_el):
                            is_el.remove(child)
                        t_el = ET.SubElement(is_el, f'{{{NS}}}t')
                        t_el.text = new_val
                        t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
                else:
                    v_el = c_el.find(f'{{{NS}}}v')
                    if v_el is not None:
                        c_el.remove(v_el)
                    c_el.set('t', 'inlineStr')
                    is_el = ET.SubElement(c_el, f'{{{NS}}}is')
                    t_el = ET.SubElement(is_el, f'{{{NS}}}t')
                    t_el.text = new_val
                    t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

    if ss_root is not None:
        for i, si in enumerate(ss_root.findall('ns:si', ns)):
            if i in index_to_translation:
                new_text = index_to_translation[i]
                for child in list(si):
                    si.remove(child)
                t_el = ET.SubElement(si, f'{{{NS}}}t')
                t_el.text = new_text
                t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')

    # 列宽
    cols_el = sheet_root.find(f'{{{NS}}}cols')
    original_col_widths = {}
    if cols_el is not None:
        for col_el in cols_el.findall(f'{{{NS}}}col'):
            min_c = int(col_el.get('min', 0))
            max_c = int(col_el.get('max', 0))
            width = float(col_el.get('width', 8))
            for c in range(min_c, max_c + 1):
                original_col_widths[c] = width

    col_max_width = defaultdict(float)
    for (row, col), val in translations.items():
        r1, c1 = row + 1, col + 1
        if (r1, c1) in skip_cells:
            continue
        text_w = _estimate_text_width(val)
        span = merged_col_spans.get((r1, c1), 1)
        per_col_w = text_w / span
        for offset in range(span):
            col_max_width[c1 + offset] = max(col_max_width[c1 + offset], per_col_w)

    new_col_widths = dict(original_col_widths)
    for col, text_w in col_max_width.items():
        original_w = original_col_widths.get(col, 8)
        needed = text_w * 1.1 + 1
        if needed > original_w * 3:
            new_col_widths[col] = min(original_w * 1.5, 60)

    if cols_el is not None:
        sheet_root.remove(cols_el)
    if new_col_widths:
        cols_el = ET.SubElement(sheet_root, f'{{{NS}}}cols')
        sheet_data = sheet_root.find(f'{{{NS}}}sheetData')
        if sheet_data is not None:
            idx = list(sheet_root).index(sheet_data)
            sheet_root.remove(cols_el)
            sheet_root.insert(idx, cols_el)
        for col in sorted(new_col_widths.keys()):
            col_el = ET.SubElement(cols_el, f'{{{NS}}}col')
            col_el.set('min', str(col))
            col_el.set('max', str(col))
            col_el.set('width', f'{new_col_widths[col]:.2f}')
            col_el.set('customWidth', '1')

    # 样式
    styles_root = ET.fromstring(styles_xml)
    cell_xfs = styles_root.find(f'{{{NS}}}cellXfs')
    if cell_xfs is not None:
        for xf in cell_xfs.findall(f'{{{NS}}}xf'):
            alignment = xf.find(f'{{{NS}}}alignment')
            if alignment is None:
                alignment = ET.SubElement(xf, f'{{{NS}}}alignment')
            alignment.set('wrapText', '1')
            if alignment.get('shrinkToFit'):
                del alignment.attrib['shrinkToFit']
            xf.set('applyAlignment', '1')

    new_shared = ET.tostring(ss_root, encoding='unicode', xml_declaration=True) if ss_root is not None else None
    new_sheet = ET.tostring(sheet_root, encoding='unicode', xml_declaration=True)
    new_styles = ET.tostring(styles_root, encoding='unicode', xml_declaration=True)

    with zipfile.ZipFile(original_path, 'r') as zin:
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.namelist():
                if item == 'xl/sharedStrings.xml' and new_shared is not None:
                    zout.writestr(item, new_shared)
                elif item == 'xl/worksheets/sheet1.xml':
                    zout.writestr(item, new_sheet)
                elif item == 'xl/styles.xml':
                    zout.writestr(item, new_styles)
                else:
                    zout.writestr(item, zin.read(item))


# ══════════════════════════════════════════════════════════════════════════════
# Streamlit 界面
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="图纸/表格翻译工具", page_icon="🌐", layout="centered")
st.title("📊 图纸 & 表格翻译工具")
st.caption("上传 Excel 表格或工程图纸 PDF，自动脱敏翻译")

# 初始化 session_state
if "result_zip" not in st.session_state:
    st.session_state.result_zip = None
    st.session_state.result_filename = None
    st.session_state.result_cost = None
    st.session_state.translating = False
if "pdf_result" not in st.session_state:
    st.session_state.pdf_result = None
    st.session_state.pdf_filename = None
    st.session_state.pdf_processing = False

# API Key 配置（从环境变量或侧边栏输入）
api_key = os.getenv("OPENAI_API") or os.getenv("OPENAI_API_KEY") or ""
baidu_api_key = os.getenv("BAIDU_API_KEY") or ""
baidu_secret_key = os.getenv("BAIDU_SECRET_KEY") or ""
with st.sidebar:
    st.header("设置")
    api_key_input = st.text_input(
        "OpenAI API Key",
        value=api_key,
        type="password",
        help="留空则使用服务器环境变量中的 Key",
    )
    if api_key_input:
        api_key = api_key_input

    st.divider()
    st.subheader("百度 OCR（图纸用）")
    baidu_key_input = st.text_input(
        "BAIDU_API_KEY",
        value=baidu_api_key,
        type="password",
    )
    baidu_secret_input = st.text_input(
        "BAIDU_SECRET_KEY",
        value=baidu_secret_key,
        type="password",
    )
    if baidu_key_input:
        baidu_api_key = baidu_key_input
    if baidu_secret_input:
        baidu_secret_key = baidu_secret_input

# 选择模式
tab_excel, tab_pdf = st.tabs(["📊 Excel 表格翻译", "📐 工程图纸脱敏翻译"])

# ══════════════════════════════════════════════════════════════════════════════
# Tab 1: Excel 翻译（原有功能）
# ══════════════════════════════════════════════════════════════════════════════

with tab_excel:
    uploaded_file = st.file_uploader("上传 Excel 文件", type=["xlsx", "xls"], key="excel_upload")
    if uploaded_file is None:
        st.session_state.result_zip = None

    lang_options = {"en": "English (英语)", "th": "Thai (泰语)", "my": "Burmese (缅甸语)"}
    selected_langs = st.multiselect(
        "目标语言",
        options=list(lang_options.keys()),
        default=["en", "th", "my"],
        format_func=lambda x: lang_options[x],
    )

    if st.button("开始翻译", type="primary", disabled=not uploaded_file or not selected_langs or st.session_state.translating, key="excel_btn"):
        if not uploaded_file or not selected_langs:
            st.stop()
        if not api_key:
            st.error("请在侧边栏输入 OpenAI API Key，或在服务器设置环境变量。")
            st.stop()

        st.session_state.translating = True
        st.session_state.result_zip = None
        client = OpenAI(api_key=api_key)

        # 保存上传文件到临时目录
        tmp_dir = tempfile.mkdtemp()
        input_path = os.path.join(tmp_dir, uploaded_file.name)
        with open(input_path, "wb") as f:
            f.write(uploaded_file.getvalue())

        ext = Path(uploaded_file.name).suffix.lower()
        stem = Path(uploaded_file.name).stem

        # xls 自动转 xlsx（保留图片）
        if ext == '.xls':
            try:
                st.info("检测到 .xls 格式，正在转换为 .xlsx（保留图片）...")
                input_path = ensure_xlsx(input_path)
            except Exception as e:
                st.error(f".xls 转换失败: {e}。请手动用 WPS/Excel 另存为 .xlsx 后再上传。")
                st.session_state.translating = False
                st.stop()

        # 读取文件
        try:
            cells, nrows, ncols, sheet_name = read_file(input_path)
        except Exception as e:
            st.error(f"读取文件失败: {e}")
            st.session_state.translating = False
            st.stop()

        translatable = {k: v for k, v in cells.items() if is_translatable(v)}
        st.info(f"共 {len(cells)} 个单元格，{len(translatable)} 个需要翻译")

        table_context = build_table_context(cells, nrows, ncols)

        total_tokens = {"input": 0, "output": 0}
        result_files = {}  # lang -> (filename, bytes)

        for lang in selected_langs:
            lang_name = LANGUAGES[lang]
            st.subheader(f"翻译为 {lang_name}")
            progress = st.progress(0)
            status = st.empty()

            # 阶段一：术语表
            status.text(f"提取 {lang_name} 术语表...")
            try:
                glossary, g_tokens = extract_glossary(client, translatable, lang)
                total_tokens["input"] += g_tokens["input"]
                total_tokens["output"] += g_tokens["output"]
            except Exception as e:
                st.error(f"术语提取失败: {e}")
                continue

            with st.expander(f"{lang_name} 术语表", expanded=False):
                st.text(glossary)

            # 阶段二：分批翻译
            keys = sorted(translatable.keys())
            batches = group_by_rows(keys, BATCH_SIZE)
            all_translations = {}

            for batch_idx, batch_keys in enumerate(batches):
                rows_in_batch = sorted(set(r for r, c in batch_keys))
                row_range = f"R{rows_in_batch[0]+1}-R{rows_in_batch[-1]+1}"
                status.text(f"批次 {batch_idx+1}/{len(batches)} ({len(batch_keys)} 格, {row_range})...")
                progress.progress((batch_idx + 1) / len(batches))

                try:
                    result, b_tokens = translate_batch(
                        client, batch_keys, translatable, table_context, glossary, lang
                    )
                    total_tokens["input"] += b_tokens["input"]
                    total_tokens["output"] += b_tokens["output"]
                    parsed = parse_translation_result(result)
                    all_translations.update(parsed)
                except Exception as e:
                    st.warning(f"批次 {batch_idx+1} 失败: {e}，重试中...")
                    try:
                        result, b_tokens = translate_batch(
                            client, batch_keys, translatable, table_context, glossary, lang
                        )
                        total_tokens["input"] += b_tokens["input"]
                        total_tokens["output"] += b_tokens["output"]
                        parsed = parse_translation_result(result)
                        all_translations.update(parsed)
                    except Exception as e2:
                        st.error(f"批次 {batch_idx+1} 再次失败: {e2}，已跳过")

            # 补上不需要翻译的单元格
            for k, v in cells.items():
                if k not in all_translations:
                    all_translations[k] = v

            # 保存结果
            output_filename = f"{stem}_{lang}.xlsx"
            output_path = os.path.join(tmp_dir, output_filename)
            try:
                save_translated_xlsx(input_path, all_translations, output_path)
                with open(output_path, "rb") as f:
                    result_files[lang] = (output_filename, f.read())
                status.text(f"{lang_name} 翻译完成！{len(all_translations)} 个单元格")
            except Exception as e:
                st.error(f"保存 {lang_name} 结果失败: {e}")

        # 打包所有结果为 zip
        if result_files:
            import io
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                for lang, (filename, data) in result_files.items():
                    zf.writestr(filename, data)
            zip_buffer.seek(0)

            st.session_state.result_zip = zip_buffer.getvalue()
            st.session_state.result_filename = f"{stem}_translated.zip"

            # 费用统计
            model = DEFAULT_MODEL
            pricing = PRICING.get(model, {"input": 0, "output": 0})
            input_cost = total_tokens["input"] / 1_000_000 * pricing["input"]
            output_cost = total_tokens["output"] / 1_000_000 * pricing["output"]
            total_cost = input_cost + output_cost
            st.session_state.result_cost = {
                "total": total_cost,
                "input_tokens": total_tokens["input"],
                "output_tokens": total_tokens["output"],
            }

        st.session_state.translating = False
        st.rerun()

    # 显示结果（持久化，不会因为点击下载而消失）
    if st.session_state.result_zip:
        st.divider()
        st.success("翻译完成！")

        if st.session_state.result_cost:
            cost = st.session_state.result_cost
            st.metric("本次费用", f"${cost['total']:.4f}",
                      help=f"Input: {cost['input_tokens']:,} tokens, Output: {cost['output_tokens']:,} tokens")

        st.download_button(
            label="📥 下载全部翻译结果 (zip)",
            data=st.session_state.result_zip,
            file_name=st.session_state.result_filename,
            mime="application/zip",
            key="excel_download",
        )


# ══════════════════════════════════════════════════════════════════════════════
# Tab 2: 工程图纸脱敏翻译
# ══════════════════════════════════════════════════════════════════════════════

with tab_pdf:
    st.caption("上传工程图纸 PDF 或 ZIP 压缩包 → 自动脱敏 + 翻译（只处理文件名含 lay/drw 的 PDF）")

    uploaded_files = st.file_uploader(
        "上传 PDF 文件或 ZIP 压缩包",
        type=["pdf", "zip"],
        accept_multiple_files=True,
        key="pdf_upload",
    )
    if not uploaded_files:
        st.session_state.pdf_result = None

    if st.button("开始处理", type="primary",
                 disabled=not uploaded_files or st.session_state.pdf_processing,
                 key="pdf_btn"):
        if not api_key:
            st.error("请在侧边栏输入 OpenAI API Key")
            st.stop()
        if not baidu_api_key or not baidu_secret_key:
            st.error("请设置环境变量 BAIDU_API_KEY 和 BAIDU_SECRET_KEY（百度 OCR）")
            st.stop()

        st.session_state.pdf_processing = True
        st.session_state.pdf_result = None

        # 导入图纸处理模块
        import sys
        gjr_path = str(Path(__file__).resolve().parent / "gjr")
        if gjr_path not in sys.path:
            sys.path.insert(0, gjr_path)
        import run_redact_v2 as redact

        # 设置 API clients
        redact.set_client(OpenAI(api_key=api_key))
        os.environ["BAIDU_API_KEY"] = baidu_api_key
        os.environ["BAIDU_SECRET_KEY"] = baidu_secret_key

        tmp_dir = tempfile.mkdtemp()
        input_dir = os.path.join(tmp_dir, "input")
        output_dir = os.path.join(tmp_dir, "output")
        os.makedirs(input_dir, exist_ok=True)
        os.makedirs(output_dir, exist_ok=True)

        # 收集所有 PDF 文件（支持直接上传 + ZIP 解压）
        pdf_paths = []
        for uf in uploaded_files:
            save_path = os.path.join(input_dir, uf.name)
            with open(save_path, "wb") as f:
                f.write(uf.getvalue())
            if uf.name.lower().endswith(".zip"):
                # 解压 ZIP
                import zipfile as _zf
                with _zf.ZipFile(save_path, 'r') as zf:
                    for name in zf.namelist():
                        if name.lower().endswith(".pdf") and not name.startswith("__MACOSX"):
                            zf.extract(name, input_dir)
                            pdf_paths.append(os.path.join(input_dir, name))
            else:
                pdf_paths.append(save_path)

        # 过滤：只处理文件名含 "lay" 或 "drw" 的 PDF（连续匹配）
        def should_process(filename):
            name_lower = filename.lower()
            return "lay" in name_lower or "drw" in name_lower

        filtered_paths = [p for p in pdf_paths if should_process(os.path.basename(p))]
        skipped = [os.path.basename(p) for p in pdf_paths if not should_process(os.path.basename(p))]

        if skipped:
            st.info(f"跳过 {len(skipped)} 个文件（不含 lay/drw）: {', '.join(skipped[:5])}")
        if not filtered_paths:
            st.warning("没有符合条件的 PDF 文件（文件名需包含 lay 或 drw）")
            st.session_state.pdf_processing = False
            st.stop()

        st.info(f"处理 {len(filtered_paths)} 个文件")
        progress = st.progress(0)
        status = st.empty()
        result_files = {}  # filename -> bytes

        for fi, input_path in enumerate(filtered_paths):
            import fitz
            import shutil
            filename = os.path.basename(input_path)
            stem = Path(filename).stem

            doc = fitz.open(input_path)
            num_pages = len(doc)
            doc.close()

            output_path = os.path.join(output_dir, filename)
            shutil.copy2(input_path, output_path)

            status.text(f"[{fi+1}/{len(filtered_paths)}] {filename} ({num_pages} 页)")

            for page_num in range(num_pages):
                pct = (fi + (page_num + 0.5) / num_pages) / len(filtered_paths)
                progress.progress(min(pct, 1.0))
                status.text(f"[{fi+1}/{len(filtered_paths)}] {filename} 第 {page_num+1}/{num_pages} 页")

                try:
                    import io, contextlib
                    log_buf = io.StringIO()
                    with contextlib.redirect_stdout(log_buf):
                        product_name, decisions = redact.process_page(
                            input_path, output_path, page_num,
                            Path(output_dir), stem,
                        )
                except Exception as e:
                    st.warning(f"{filename} 第 {page_num+1} 页出错: {e}")

            # 压缩
            _doc = fitz.open(output_path)
            _doc.save(output_path + ".tmp", garbage=4, deflate=True)
            _doc.close()
            shutil.move(output_path + ".tmp", output_path)

            with open(output_path, "rb") as f:
                result_files[filename] = f.read()

        progress.progress(1.0)
        status.text("完成！")

        # 打包结果
        if len(result_files) == 1:
            fname, data = next(iter(result_files.items()))
            st.session_state.pdf_result = data
            st.session_state.pdf_filename = fname
        else:
            # 多文件打包 ZIP（按原始顺序排列）
            import io as _io
            zip_buf = _io.BytesIO()
            import zipfile as _zf2
            # 按 filtered_paths 的原始顺序写入
            ordered_names = [os.path.basename(p) for p in filtered_paths]
            with _zf2.ZipFile(zip_buf, 'w', _zf2.ZIP_DEFLATED) as zf:
                for fname in ordered_names:
                    if fname in result_files:
                        zf.writestr(fname, result_files[fname])
            zip_buf.seek(0)
            st.session_state.pdf_result = zip_buf.getvalue()
            # 使用原始上传的 ZIP 文件名，或首个文件名
            orig_zip_name = None
            for uf in uploaded_files:
                if uf.name.lower().endswith(".zip"):
                    orig_zip_name = uf.name
                    break
            st.session_state.pdf_filename = orig_zip_name or f"{Path(ordered_names[0]).stem}_batch.zip"

        st.session_state.pdf_product = f"{len(result_files)} 个文件"
        st.session_state.pdf_processing = False
        st.rerun()

    # 显示结果
    if st.session_state.pdf_result:
        st.divider()
        st.success(f"处理完成！{st.session_state.get('pdf_product', '')}")
        mime = "application/pdf" if st.session_state.pdf_filename.endswith(".pdf") else "application/zip"
        st.download_button(
            label=f"📥 下载 {st.session_state.pdf_filename}",
            data=st.session_state.pdf_result,
            file_name=st.session_state.pdf_filename,
            mime=mime,
            key="pdf_download",
        )
