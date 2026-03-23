import os
import re
import base64
import time
import sys
import requests
from dotenv import load_dotenv

load_dotenv()

API_KEY = os.getenv("BAIDU_API_KEY")
SECRET_KEY = os.getenv("BAIDU_SECRET_KEY")


def md_to_xlsx(md_content, output_path):
    """从 markdown 中提取所有 HTML 表格，转为 xlsx"""
    try:
        import openpyxl
    except ImportError:
        print("需要 openpyxl: pip install openpyxl")
        return

    tables = re.findall(r'<table\b[^>]*>.*?</table>', md_content, re.DOTALL)
    if not tables:
        print("未找到表格")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for idx, table_html in enumerate(tables):
        ws = wb.create_sheet(title=f"Table_{idx+1}")
        rows = re.findall(r'<tr[^>]*>(.*?)</tr>', table_html, re.DOTALL)
        for r, row in enumerate(rows, 1):
            cells = re.findall(r'<td[^>]*>(.*?)</td>', row, re.DOTALL)
            for c, cell in enumerate(cells, 1):
                ws.cell(row=r, column=c, value=re.sub(r'<[^>]+>', '', cell).strip())

    wb.save(output_path)
    print(f"已保存到 {output_path}")


def get_access_token():
    url = "https://aip.baidubce.com/oauth/2.0/token"
    resp = requests.post(url, params={
        "grant_type": "client_credentials",
        "client_id": API_KEY,
        "client_secret": SECRET_KEY
    })
    data = resp.json()
    if "access_token" in data:
        print(f"access_token 获取成功")
        return data["access_token"]
    else:
        print(f"获取 token 失败: {data}")
        sys.exit(1)


def submit_task(access_token, file_path):
    url = f"https://aip.baidubce.com/rest/2.0/brain/online/v2/paddle-vl-parser/task?access_token={access_token}"

    with open(file_path, "rb") as f:
        file_data = base64.b64encode(f.read()).decode()

    resp = requests.post(url, data={
        "file_data": file_data,
        "file_name": os.path.basename(file_path)
    }, timeout=300)

    data = resp.json()
    print(f"提交响应: {data}")

    if "result" in data and "task_id" in data["result"]:
        return data["result"]["task_id"]
    else:
        print(f"提交失败: {data}")
        sys.exit(1)


def query_result(access_token, task_id):
    url = f"https://aip.baidubce.com/rest/2.0/brain/online/v2/paddle-vl-parser/task/query?access_token={access_token}"

    for i in range(20):
        time.sleep(5)
        resp = requests.post(url, data={"task_id": task_id})
        data = resp.json()
        status = data.get("result", {}).get("status")
        print(f"轮询 {i+1}: status={status}")

        if status == "success":
            result = data["result"]
            md_url = result.get("markdown_url")
            json_url = result.get("parse_result_url")
            print(f"\nMarkdown 结果: {md_url}")
            print(f"JSON 结果: {json_url}")

            output_name = os.path.splitext(os.path.basename(file_path))[0]

            # 下载 markdown 结果
            if md_url:
                md_content = requests.get(md_url).text
                with open(f"output/{output_name}_baidu.md", "w", encoding="utf-8") as f:
                    f.write(md_content)
                print(f"已保存到 output/{output_name}_baidu.md")

                # 从 markdown 中提取 HTML 表格转 xlsx
                md_to_xlsx(md_content, f"output/{output_name}_baidu.xlsx")

            # 下载 JSON 结果
            if json_url:
                json_content = requests.get(json_url).text
                with open(f"output/{output_name}_baidu.json", "w", encoding="utf-8") as f:
                    f.write(json_content)
                print(f"已保存到 output/{output_name}_baidu.json")

            return data

        elif status == "failed":
            print(f"任务失败: {data}")
            return data

    print("轮询超时")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python test_baidu_api.py <pdf文件路径>")
        sys.exit(1)

    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"文件不存在: {file_path}")
        sys.exit(1)

    os.makedirs("output", exist_ok=True)

    token = get_access_token()
    task_id = submit_task(token, file_path)
    print(f"task_id: {task_id}")
    query_result(token, task_id)
